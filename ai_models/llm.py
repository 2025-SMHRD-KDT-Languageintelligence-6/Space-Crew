import os
import re
import shutil
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException
from faster_whisper import WhisperModel
from transformers import pipeline
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# 1. 초기 설정
app = FastAPI(title="Space-PMS AI Module (Team Standard)")

DB_CONFIG = {
}

@app.get("/")
async def health_check():
    return {
        "status": "ok",
        "message": "PCPMS AI Server is live!",
        "port": 8001,
        "timestamp": datetime.now().isoformat()
    }

# --- [모델 로딩] ---
# 팀원들과 공유 시 첫 실행 때 모델 다운로드 시간이 소요됨을 공지하세요.
print("🚀 AI 모델 로딩 중 (CPU 모드 / Port: 8001)...")

# STT: Faster-Whisper
stt_model = WhisperModel("base", device="cpu", compute_type="int8")

# 요약: KoBART
summarizer = pipeline(
    "summarization",
    model="digit82/kobart-summarization",
    device=-1  # CPU 강제 사용
)
print("✅ 모델 로딩 완료!")

# 파일 저장 경로 설정 (상대 경로 사용으로 팀원간 호환성 확보)
# BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
# RESULT_DIR = os.path.join(BASE_DIR, "results")
BASE_STORAGE = r"D:\pms_uploads"
UPLOAD_DIR = os.path.join(BASE_STORAGE, "uploads")
RESULT_DIR = os.path.join(BASE_STORAGE, "results") 

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(RESULT_DIR, exist_ok=True)

# 2. 유틸리티 함수
def extract_action_items(text):
    """규칙 기반 결정사항 추출"""
    keywords = [r"해야 함", r"하기로 함", r"결정되었습니다", r"필요함", r"진행 예정", r"담당:"]
    # 문장 분리 로직 보강
    sentences = re.split(r'\.|\n', text)
    found = [s.strip() for s in sentences if any(re.search(kw, s) for kw in keywords)]
    return "\n".join(found) if found else "특이사항 없음"

def save_to_excel(data, output_path):
    """엑셀 양식 저장"""
    wb = Workbook()
    ws = wb.active
    ws.title = "회의록"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    left_center_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 헤더
    ws.merge_cells('A1:D2')
    ws['A1'] = "회 의 록"
    ws['A1'].font = Font(name='맑은 고딕', size=18, bold=True)
    ws['A1'].alignment = center_align

    # 작성자/제목
    ws.cell(row=3, column=1, value="작 성 자").alignment = center_align
    ws.cell(row=3, column=2, value=data['writer']).alignment = left_center_align
    ws.cell(row=4, column=1, value="제  목").alignment = center_align
    ws.cell(row=4, column=2, value=data['title']).alignment = left_center_align

    # 요약
    ws.cell(row=5, column=1, value="요  약").alignment = center_align
    ws.merge_cells('B5:D5')
    ws.cell(row=5, column=2, value=data['summary']).alignment = left_center_align
    ws.row_dimensions[5].height = 100

    # 결정사항
    ws.cell(row=6, column=1, value="결정사항").alignment = center_align
    ws.merge_cells('B6:D6')
    ws.cell(row=6, column=2, value=data['action_items']).alignment = left_center_align
    ws.row_dimensions[6].height = 100

    # 테두리 일괄 적용
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    wb.save(output_path)

# 3. API 엔드포인트
@app.post("/process-meeting")

async def process_meeting(file: UploadFile = File(...)):
    print("✅ /process-meeting called:", file.filename)
    try:
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"{timestamp}_{file.filename}"
        file_path = os.path.join(UPLOAD_DIR, safe_filename)

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # (1) STT 실행
        segments, _ = stt_model.transcribe(file_path, beam_size=5, language="ko")
        full_text = " ".join([s.text for s in segments])

        # (2) 요약 실행
        if len(full_text.strip()) > 20:
            summary_res = summarizer(full_text[:1000], max_length=150, min_length=30, do_sample=False)
            summary = summary_res[0]['summary_text']
        else:
            summary = "분석할 내용이 충분하지 않습니다."

        # (3) 결과 조립 및 엑셀 생성
        action_items = extract_action_items(full_text)
        result_data = {
            "title": f"회의분석_{timestamp}",
            "writer": "Space-PMS AI",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "full_text": full_text,
            "summary": summary,
            "action_items": action_items
        }

        excel_name = f"result_{timestamp}.xlsx"
        excel_path = os.path.join(RESULT_DIR, excel_name)
        save_to_excel(result_data, excel_path)

        # import pymysql
        # # DB_CONFIG는 파일 최상단 혹은 함수 외부에 정의되어 있어야 합니다!
        # conn = pymysql.connect(**DB_CONFIG)
        # try:
        #     with conn.cursor() as cursor:
        #         sql = """
        #             INSERT INTO TB_MEETING (
        #                 MEET_TITLE, MEET_DT, CONTENT_FULL, CONTENT_SUM, ACTION_ITEMS, FILE_PATH, LAST_UPDUSR_ID
        #             ) VALUES (%s, NOW(), %s, %s, %s, %s, %s)
        #         """
        #         cursor.execute(sql, (
        #             result_data['title'],
        #             result_data['full_text'],
        #             result_data['summary'],
        #             result_data['action_items'],
        #             excel_path,  # 엑셀 경로
        #             "AI_SYSTEM"
        #         ))
        #     conn.commit()
        #     print("✅ [DB] 회의록 분석 데이터 저장 완료!")
        # except Exception as db_e:
        #     print(f"❌ [DB] 저장 중 에러 발생: {db_e}")
        #     # DB 저장이 실패해도 일단 분석 결과는 보여주려면 pass, 
        #     # 엄격하게 하려면 raise HTTPException
        # finally:
        #     conn.close()

        return {
            "status": "success",
            "data": result_data,
            "excel_path": excel_path.replace("\\", "/") # 경로 통일
        }

    except Exception as e:
        print(f"❌ 서버 에러 발생: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    # 팀 표준 포트 8001로 실행
    uvicorn.run(app, host="127.0.0.1", port=8001)